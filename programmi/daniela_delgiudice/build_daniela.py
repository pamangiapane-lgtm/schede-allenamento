#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Genera Sedute_Daniela_DelGiudice.xlsx — 8 sedute esplose (4 sett x 2/sett)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/user/schede-allenamento/programmi/daniela_delgiudice/Sedute_Daniela_DelGiudice.xlsx"

# ---- palette ----
WK = {1: ("W1 · Fitness posturale", "BDD7EE"),
      2: ("W2 · Controllo motorio", "C6E0B4"),
      3: ("W3 · Transizione al fitness", "F8CBAD"),
      4: ("W4 · Fitness a basso rischio", "FFE699")}
DARK = "1F4E79"; MID = "2E75B6"; GREY = "808080"; REDF = "C00000"
WHITE = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

def fill(hexv): return PatternFill("solid", fgColor=hexv)

wb = openpyxl.Workbook()

# ============================================================ PANORAMICA
ws = wb.active; ws.title = "Panoramica"
ws.sheet_view.showGridLines = False
ws["A1"] = "DANIELA DEL GIUDICE (53) — Sedute in sala con il coach"
ws["A1"].font = Font(size=15, bold=True, color=DARK)
ws["A2"] = "8 sedute · 2/settimana · 45–50′ · progressione 4 settimane · binario COMPLEMENTARE al recupero funzionale"
ws["A2"].font = Font(size=10, italic=True, color=GREY)
ws.merge_cells("A1:H1"); ws.merge_cells("A2:H2")

ws["A4"] = "⚠️  Quadro clinico: cervicoartrosi C2–C7 con ernie/stenosi C6–C7 · flogosi scapolo-omerale sx · ipertesa (Olpress)."
ws["A4"].font = Font(bold=True, color=REDF); ws.merge_cells("A4:H4")
ws["A5"] = "Svolgere con via libera medico/fisio. NON sostituisce il recupero funzionale (che Daniela fa a casa). Sintomi → stop."
ws["A5"].font = Font(italic=True, color=REDF); ws.merge_cells("A5:H5")

# mappa settimane
r = 7
ws.cell(r,1,"MAPPA DELLE 4 SETTIMANE").font = Font(bold=True, color=DARK)
r += 1
hdr = ["Settimana","Focus","Carico / RPE","Idea guida"]
for c,h in enumerate(hdr,1):
    cell = ws.cell(r,c,h); cell.fill = fill(MID); cell.font = WHITE; cell.alignment = CTR; cell.border = BORD
rows = [
 ("W1","Fitness posturale / rieducazione","c.libero · RPE 3–4","Consapevolezza, respiro, tecnica pulita, fiducia nel movimento"),
 ("W2","Controllo motorio","carichi leggerissimi · RPE 4–5","Pattern consolidati, propriocezione, attivazione glutei/scapole"),
 ("W3","Transizione al fitness","leggero · RPE 5–6","+volume, carries, hip thrust, equilibrio libero"),
 ("W4","Fitness a basso rischio consolidato","leggero-moderato · RPE 5–6","Circuito sostenibile + ri-verifica marker"),
]
for i,row in enumerate(rows):
    rr = r+1+i
    for c,val in enumerate(row,1):
        cell = ws.cell(rr,c,val); cell.border = BORD; cell.alignment = LFT if c>1 else CTR
    ws.cell(rr,1).fill = fill(WK[i+1][1])

# calendario sedute
r = r+len(rows)+2
ws.cell(r,1,"LE 8 SEDUTE").font = Font(bold=True, color=DARK); r+=1
for c,h in enumerate(["Seduta","Settimana","Tipo","Focus della seduta"],1):
    cell = ws.cell(r,c,h); cell.fill = fill(MID); cell.font = WHITE; cell.alignment = CTR; cell.border = BORD
sess = [
 ("S1",1,"A","Gambe / catena posteriore + toracica"),
 ("S2",1,"B","Anca / glutei + core + equilibrio + spalla"),
 ("S3",2,"A","Gambe / catena posteriore + toracica"),
 ("S4",2,"B","Anca / glutei + core + equilibrio + spalla"),
 ("S5",3,"A","Gambe / catena posteriore + toracica"),
 ("S6",3,"B","Anca / glutei + core + equilibrio + spalla"),
 ("S7",4,"A","Gambe / catena posteriore + toracica"),
 ("S8",4,"B","Anca / glutei + core + equilibrio + spalla"),
]
for i,(sid,wk,tp,foc) in enumerate(sess):
    rr = r+1+i
    for c,val in enumerate([sid,WK[wk][0].split(' · ')[0],tp,foc],1):
        cell = ws.cell(rr,c,val); cell.border = BORD; cell.alignment = LFT if c==4 else CTR
    ws.cell(rr,2).fill = fill(WK[wk][1])

for col,w in {"A":11,"B":13,"C":22,"D":46,"E":10,"F":16,"G":30,"H":30}.items():
    ws.column_dimensions[col].width = w

# ============================================================ LEGENDA & VINCOLI
lg = wb.create_sheet("Legenda & Vincoli")
lg.sheet_view.showGridLines = False
lg["A1"]="LEGENDA, VINCOLI E SICUREZZA"; lg["A1"].font=Font(size=14,bold=True,color=DARK); lg.merge_cells("A1:D1")
blocks = [
 ("RPE (percezione dello sforzo)", MID, [
   "Scala 1–10 (1 = passeggiata, 10 = massimo). Per Daniela restiamo SEMPRE ≤ 6.",
   "W1 RPE 3–4 · W2 RPE 4–5 · W3–W4 RPE 5–6. Deve poter parlare e respirare fluida (no apnea).",
 ]),
 ("Carichi indicativi", MID, [
   "I kg suggeriti sono un PUNTO DI PARTENZA CONSERVATIVO: da tarare sulla 1ª seduta.",
   "Priorità assoluta: tecnica pulita + ASSENZA DI SINTOMI, non il numero. Nel dubbio, meno carico.",
   "Si sale di carico SOLO se la seduta precedente è pulita e senza sintomi nelle 24h (gate).",
 ]),
 ("❌ VIETATO SEMPRE (dalla relazione)", REDF, [
   "Bilanciere sul collo / carico assiale sul rachide (no back squat, no military).",
   "Overhead con carico in stazione eretta (trigger vertigini).",
   "Addominali in flessione (sit-up, crunch).",
   "Stretching statico passivo forzato del collo.",
   "Valsalva / apnea sotto sforzo (è ipertesa cronica).",
   "Impatto/salti e cambi rapidi di posizione del capo.",
 ]),
 ("🛑 STOP e consulto medico (red flags)", REDF, [
   "Parestesie / formicolii / dolore irradiato al braccio; perdita di forza o presa alla mano.",
   "Vertigini, nistagmo, sbandamenti; deficit visivi.",
   "Picchi pressori o dolore toracico.",
 ]),
 ("↩️ REGREDIRE di un gradino se", "ED7D31", [
   "Dolore cervicale/spalla sx VAS > 4 che persiste oltre 24h.",
   "Sbandamenti/fastidi posizione-dipendenti durante gli esercizi.",
   "Affaticamento precoce che rompe la neutralità delle curve.",
 ]),
]
rr = 3
for title,color,lines in blocks:
    c = lg.cell(rr,1,title); c.fill=fill(color); c.font=WHITE; c.alignment=LFT; lg.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=4)
    rr+=1
    for ln in lines:
        cc = lg.cell(rr,1,"•  "+ln); cc.alignment=LFT; lg.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=4); rr+=1
    rr+=1
lg.column_dimensions["A"].width=120

# ============================================================ SEDUTE
COLS = ["Blocco","Esercizio","Attrezzo","Serie × Reps/Durata","RPE","Carico indicativo",
        "Sicurezza / \"guarda che…\"","Note & osservazioni (compilare in seduta)"]
CW = [24,34,20,20,7,20,40,34]

def seduta_A(wk):
    L = {1:dict(rpe="3–4",gob="box / c.libero",rdl="bastone o 2 kg",row="leggero",carry="KB 6–8 kg",ht=None,cond="Bike facile 5–6′"),
         2:dict(rpe="4–5",gob="KB 6–8 kg",rdl="DB 2×4–5 kg",row="leggero +",carry="KB 8–10 kg",ht=None,cond="Bike facile 6–8′"),
         3:dict(rpe="5–6",gob="KB 8–12 kg",rdl="DB 2×5–7 kg",row="moderato",carry="KB 10–12 kg",ht="DB/bilanc. leggero",cond="Bike/sled 8–10′"),
         4:dict(rpe="5–6",gob="KB 8–12 kg",rdl="DB 2×5–7 kg",row="moderato",carry="KB 10–12 kg",ht="DB/bilanc. leggero",cond="Bike/sled 8–10′")}[wk]
    rows=[
     ("1. Reset & respiro + mobilità (10′)","90/90 breathing (gambe su rialzo)","Box/tappetino","3 × 5 respiri","—","—","Espira lungo, costato giù, lombare appoggiata (riduce l'anteversione)"),
     ("","Estensione toracica su foam roller (T3–T5)","Foam roller","2 × 8","—","c.libero","Muovi la TORACICA, non il collo; testa sostenuta"),
     ("","Allungo attivo flessori d'anca (half-kneeling)","Tappetino","2 × 30″/lato","3","c.libero","Glutei attivi, NON inarcare la lombare"),
    ]
    if wk>=3:
        rows.append(("","Thread the needle (rotazione toracica quadrupedia)","Tappetino","2 × 6/lato","—","c.libero","Ruota il torace, bacino fermo, collo rilassato"))
    rows += [
     ("2. Controllo motorio & attivazione (10–12′)",
        "Box squat (schema)" if wk==1 else "Goblet squat al box (schema)","Box (+KB)","3 × 8","3–4","c.libero" if wk==1 else "KB leggero","Peso sui talloni, tronco eretto, ginocchia in linea"),
     ("","Hip hinge con bastone (3 punti di contatto)" if wk==1 else "Hip hinge DB (schema RDL)","Bastone / DB","3 × 8","3–4","c.libero","Bacino indietro, schiena lunga, RACHIDE NEUTRO, no flessione del collo"),
     ("","Equilibrio monopodalico "+("(con appoggio)" if wk<=2 else "(libero + passaggio oggetto)"),"—/rack","3 × 20–30″/lato","—","—","Sguardo avanti, NIENTE scatti del capo; progredisci togliendo l'appoggio"),
    ]
    if wk>=2:
        rows.append(("","Monster walk mini-band (attivazione glutei)","Skimmy","2 × 10/direzione","3","band leggera","Ginocchia larghe, bacino stabile, passi controllati"))
    rows += [
     ("3. Fitness a basso rischio (15–18′)","Goblet squat al box","KB","3 × "+("8" if wk<3 else "10"),L["rpe"],L["gob"],"Tronco eretto, NO apnea; carico anteriore ≠ carico sul collo"),
     ("","Hip hinge / RDL","DB o bilanciere leggero","3 × 8",L["rpe"],L["rdl"],"Schiena lunga, no flessione del collo; senti dietro la coscia"),
     ("","Row orizzontale (lat machine seduta o DB con appoggio petto)","Lat machine / DB","3 × 12",L["rpe"],L["row"],"Scapole che scorrono, gomiti bassi, NO shrug, collo neutro"),
    ]
    if L["ht"]:
        rows.append(("","Hip thrust (pad su anche)","Bilanciere/DB","3 × 10",L["rpe"],L["ht"],"Spinta dai talloni, coste giù; NON iper-estendere la lombare; mento leggermente retratto"))
    rows += [
     ("","Suitcase carry (un lato)","KB / manubrio","3 × 20 m/lato","≤5",L["carry"],"Busto DRITTO e fermo (anti-inclinazione), spalle basse"),
     ("4. Condizionamento + scarico (8–10′)","Bike verticale o sled leggero (spinta bassa)","Bike / sled",L["cond"],"≤5","facile","Collo neutro, respiro fluido, NO Valsalva"),
     ("","Scarico: respiro diaframmatico + estensione toracica dolce","Tappetino","2 × 6 resp","—","—","Chiudere in decompressione"),
    ]
    return rows

def seduta_B(wk):
    L = {1:dict(rpe="3–4",split="c.libero",sldl="c.libero/appoggio",step="c.libero",er="elastico soft",cond="Camminata salita 5–6′"),
         2:dict(rpe="4–5",split="DB 2×3–4 kg",sldl="KB 4–6 kg",step="DB 2×3–4 kg",er="elastico soft",cond="Camminata/bike 6–8′"),
         3:dict(rpe="5–6",split="DB 2×5–6 kg",sldl="KB 6–8 kg",step="DB 2×5–6 kg",er="elastico medio",cond="Intervalli bike facili 8–10′"),
         4:dict(rpe="5–6",split="DB 2×5–6 kg",sldl="KB 6–8 kg",step="DB 2×5–6 kg",er="elastico medio",cond="Intervalli bike facili 8–10′")}[wk]
    rows=[
     ("1. Reset & respiro + mobilità (10′)","Respirazione diaframmatica (ginocchia su box)","Box/tappetino","3 × 6 respiri","—","—","4″ inspira / 6″ espira, NO apnea"),
     ("","Open book (rotazione toracica sul fianco)","Tappetino","2 × 8/lato","—","c.libero","Ruota il TORACE, bacino fermo, collo rilassato"),
     ("","Mobilità anca 90/90 (transizioni lente)","Tappetino","2 × 8","—","c.libero","Cerca la rotazione interna sx SENZA forzare (rigidità capsulare nota)"),
    ]
    rows += [
     ("2. Controllo motorio & attivazione (10–12′)","Split squat assistito","Rack/appoggio (+DB)","3 × 8/gamba",L["rpe"],L["split"],"Busto eretto, ginocchio stabile (non cade dentro)"),
     ("","Ponte glutei "+("(bipodalico)" if wk==1 else "(bi → mono a fine ciclo)")+" con mini-band","Skimmy","3 × 10","3–4","band leggera","Spinta dai talloni, stringi i glutei, NON iper-estendere la lombare"),
     ("","Wall slide BASSO / scapular con elastico (sotto la spalla)","Elastico","3 × 10","≤4","soft","Serrato attivo; MAI salire in overhead"),
    ]
    if wk>=2:
        rows.append(("","Lateral band walk (skimmy)","Skimmy","2 × 10/lato","3","band leggera","Passi laterali controllati, bacino stabile"))
    rows += [
     ("3. Fitness a basso rischio (15–18′)","Step-up basso","Box basso (+DB)","3 × 8/gamba",L["rpe"],L["step"],"Spingi di gamba salendo, discesa controllata"),
     ("","Stacco monopodalico assistito (SLDL)","KB + appoggio","3 × 6/gamba",L["rpe"],L["sldl"],"Anca–ginocchio allineati, bacino livellato, schiena neutra"),
     ("","Pallof press (anti-rotazione)","Lat machine / elastico","3 × 10/lato","≤6","elastico/cavo","Bacino e coste FERME, non ruotare, respiro fluido"),
     ("","Dead bug con pullover elastico (leva corta)","Elastico","3 × 8/lato","≤5","elastico soft","LOMBARE a contatto, scendi lento il tallone"),
     ("","Extrarotazione spalla SX (gomito al fianco, 0–30°)","Elastico","3 × 12","≤5",L["er"],"RANGE PROTETTO, no dolore; lato in flogosi → gentile"),
     ("4. Condizionamento + scarico (8–10′)","Camminata in salita / intervalli bike facili","Tappeto / bike",L["cond"],"≤5","facile","Nessun trigger vertigini, respiro fluido"),
     ("","Face pull leggero (gomiti sotto la spalla)","Lat machine / elastico","2 × 15","≤5","leggero","Postura alto dorso, no shrug"),
     ("","Scarico: respiro + open book dolce","Tappetino","2 × 6 resp","—","—","Chiudere in decompressione"),
    ]
    return rows

def add_seduta(idx, wk, tipo, focus, rows):
    ws = wb.create_sheet(f"S{idx}")
    ws.sheet_view.showGridLines = False
    wname, wcol = WK[wk]
    ws.cell(1,1,f"SEDUTA {idx}  ·  {wname}  ·  Tipo {tipo}").font = Font(size=13,bold=True,color=DARK)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    b = ws.cell(2,1,f"Focus: {focus}   |   Durata 45–50′   |   RPE target {'3–4' if wk==1 else '4–5' if wk==2 else '5–6'}")
    b.font = Font(italic=True, color=GREY); ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))
    tag = ws.cell(3,1,f"Settimana {wk} — {wname.split(' · ')[1]}"); tag.fill=fill(wcol); tag.alignment=LFT
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(COLS))
    hr = 5
    for c,h in enumerate(COLS,1):
        cell = ws.cell(hr,c,h); cell.fill=fill(DARK); cell.font=WHITE; cell.alignment=CTR; cell.border=BORD
    cur = hr+1
    for row in rows:
        row = list(row) + [""]*(len(COLS)-len(row))  # pad note col
        is_block = bool(row[0])
        for c,val in enumerate(row,1):
            cell = ws.cell(cur,c,val)
            cell.border = BORD
            cell.alignment = LFT if c in (2,7,8) else CTR
            if c==1 and is_block:
                cell.font = Font(bold=True, color=DARK); cell.fill = fill(wcol)
        cur += 1
    for c,w in enumerate(CW,1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A6"

for idx,(sid,wk,tp,foc) in enumerate(sess,1):
    rows = seduta_A(wk) if tp=="A" else seduta_B(wk)
    add_seduta(idx, wk, tp, foc, rows)

# ============================================================ MONITORAGGIO & MARKER
mo = wb.create_sheet("Monitoraggio & Marker")
mo.sheet_view.showGridLines = False
mo["A1"]="MONITORAGGIO SEDUTE"; mo["A1"].font=Font(size=14,bold=True,color=DARK); mo.merge_cells("A1:I1")
mh = ["Data","Seduta","RPE medio","Sintomi collo/spalla (0–10)","Sbandamenti (S/N)","Pressione","Sonno (1–5)","Carichi usati","Note"]
for c,h in enumerate(mh,1):
    cell=mo.cell(3,c,h); cell.fill=fill(MID); cell.font=WHITE; cell.alignment=CTR; cell.border=BORD
for r in range(4,4+16):
    for c in range(1,len(mh)+1):
        mo.cell(r,c,"").border=BORD
mw=[12,9,10,16,14,12,10,22,34]
for c,w in enumerate(mw,1): mo.column_dimensions[get_column_letter(c)].width=w

# marker table
sr = 4+16+2
mo.cell(sr,1,"MARKER — inizio vs fine 4 settimane").font=Font(bold=True,color=DARK); mo.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=4)
sr+=1
for c,h in enumerate(["Marker","Iniziale","Fine 4 sett","Note"],1):
    cell=mo.cell(sr,c,h); cell.fill=fill(MID); cell.font=WHITE; cell.alignment=CTR; cell.border=BORD
markers=["Distanza occipite–muro (forward head)","Rotazione toracica (gradi indicativi)",
 "Rotazione interna anca sx (ampiezza, no dolore)","Equilibrio monopodalico DX (sec)",
 "Equilibrio monopodalico SX (sec)","Tenuta core anti-rotazione (qualità)",
 "RPE medio settimanale","Pressione a riposo"]
for i,m in enumerate(markers):
    rr=sr+1+i
    for c in range(1,5):
        cell=mo.cell(rr,c, m if c==1 else ""); cell.border=BORD; cell.alignment=LFT if c in (1,4) else CTR

wb.save(OUT)
print("OK ->", OUT, "| fogli:", wb.sheetnames)
