# -*- coding: utf-8 -*-
import datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---------- stili ----------
NAVY="0B4A6F"; WHITE="FFFFFF"
def fill(c): return PatternFill("solid", fgColor=c)
PH = {  # colori fase
 "Accumulo":"C6E0B4", "Deload":"FFE699", "Sviluppo":"F8CBAD",
 "Realizzazione":"BDD7EE", "InSeason":"D9D9D9"
}
thin = Side(style="thin", color="B9C6CE")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
hdr_font = Font(bold=True, color=WHITE, size=10)
hdr_fill = fill(NAVY)
title_font = Font(bold=True, color=NAVY, size=14)
sub_font = Font(bold=True, color=NAVY, size=11)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_header(ws, row, ncols, start=1):
    for c in range(start, start+ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border

def put(ws, r, c, val, font=None, fillc=None, al=wrap, bd=True):
    cell = ws.cell(row=r, column=c, value=val)
    if font: cell.font = font
    if fillc: cell.fill = fill(fillc)
    if al: cell.alignment = al
    if bd: cell.border = border
    return cell

def widths(ws, d):
    for col,w in d.items(): ws.column_dimensions[col].width = w

# ---------- calendario ----------
# 8 settimane Lun-Dom da Lun 31 ago 2026; poi settimana in-season con gara Sab 31 ott
start = dt.date(2026,8,31)
weeks=[]
for i in range(9):
    ms = start + dt.timedelta(days=7*i)
    weeks.append((ms, ms+dt.timedelta(days=6)))
# fase per settimana (indice 0..7 = S1..S8 ; 8 = in-season)
fase_of = ["Accumulo","Accumulo","Accumulo","Deload","Sviluppo","Sviluppo","Sviluppo","Realizzazione","InSeason"]
fase_label = {"Accumulo":"FASE 1 · Accumulo/Fondamenta","Deload":"Scarico (Deload)",
              "Sviluppo":"FASE 2 · Sviluppo/Trasferimento","Realizzazione":"FASE 3 · Realizzazione/Test",
              "InSeason":"In-season · Settimana tipo"}
def fmt(d): return d.strftime("%d/%m")

# reference 1RM (assunzione dichiarata)
RM = {"Squat":55, "Panca":30, "RDL":65}
def kg(p, oneRM):
    v = p/100.0*oneRM
    return round(v/2.5)*2.5

# ============================================================= PANORAMICA
ws = wb.active; ws.title = "Panoramica"
widths(ws, {"A":4,"B":16,"C":13,"D":13,"E":13,"F":13,"G":13,"H":13,"I":13,"J":13})
put(ws,1,2,"PREPARAZIONE PRE-SEASON — Pallavolo femminile Serie D", title_font, bd=False)
put(ws,2,2,"Documento operativo · assistente del preparatore · v0.2 (8 settimane)", Font(italic=True, color="666666"), bd=False)
r=4
put(ws,r,2,"CONTESTO SQUADRA", sub_font, bd=False); r+=1
ctx=[
 ("Categoria","Serie D femminile · atlete 18–32 anni"),
 ("Profili","Eterogenei: forma variabile; alcune equilibrate, altre con problemi motori → 3 fasce (Rosso/Giallo/Verde)"),
 ("Frequenza","3 sedute/settimana · 2 Campo + 1 Pesi · durata media 60'"),
 ("Attrezzatura campo","Elastici, skimmy (mini-band), TRX, ostacoli, coni/cinesini, box"),
 ("Attrezzatura pesi","Palestra completa"),
 ("Obiettivo","Sviluppare la condizione con progressioni proporzionali e crescenti; forza + potenza sport-specifica + prevenzione"),
 ("Periodo","Preparazione 01/09 → 25/10 · 1ª gara sab 31/10"),
]
for k,v in ctx:
    put(ws,r,2,k,Font(bold=True)); ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=10); put(ws,r,3,v); r+=1

r+=1
put(ws,r,2,"MACRO-FASI E DURATE", sub_font, bd=False); r+=1
head=["Fase","Settimane","Durata","Obiettivo dominante","Focus carico"]
for i,h in enumerate(head): put(ws,r,2+i,h)
style_header(ws,r,len(head),start=2)
r+=1
fasi=[
 ("FASE 1 · Accumulo/Fondamenta","S1–S3","3 sett.","Valutazione, qualità del movimento, forza generale, capacità di lavoro","Vol ↑ · Int bassa-media","Accumulo"),
 ("Scarico (Deload)","S4","1 sett.","Recupero/riorganizzazione; re-check fasce","Vol −40% · Int mantenuta","Deload"),
 ("FASE 2 · Sviluppo/Trasferimento","S5–S7","3 sett.","Forza (intensità ↑), pliometria, condizionamento specifico, potenza","Int ↑ · Vol medio","Sviluppo"),
 ("FASE 3 · Realizzazione/Test","S8","1 sett.","Taper, qualità/velocità, re-test; pronte alla stagione","Int alta · Vol basso","Realizzazione"),
]
for nome,sett,dur,ob,fo,col in fasi:
    put(ws,r,2,nome,Font(bold=True),PH[col]); put(ws,r,3,sett,None,PH[col],center); put(ws,r,4,dur,None,PH[col],center)
    put(ws,r,5,ob); put(ws,r,6,fo); r+=1

r+=1
put(ws,r,2,"CALENDARIO E DURATA FASI (Gantt)", sub_font, bd=False); r+=1
# intestazione settimane
put(ws,r,2,"Settimana")
for i,(ms,me) in enumerate(weeks):
    lab = f"S{i+1}\n{fmt(ms)}–{fmt(me)}" if i<8 else f"In-season\n{fmt(ms)}–{fmt(me)}"
    put(ws,r,3+i,lab)
style_header(ws,r,1+len(weeks),start=2)
r+=1
# riga fase (barra colorata)
put(ws,r,2,"Fase",Font(bold=True))
for i in range(len(weeks)):
    f=fase_of[i]; put(ws,r,3+i,{"Accumulo":"ACC","Deload":"DELOAD","Sviluppo":"SVIL","Realizzazione":"TEST","InSeason":"GARA"}[f],Font(bold=True,size=9),PH[f],center)
r+=1
# riga carico (indicativo 1-10)
carico=[5,6,7,4,7,8,8,5,6]
put(ws,r,2,"Carico rel. (1-10)",Font(bold=True))
for i,v in enumerate(carico): put(ws,r,3+i,v,None,PH[fase_of[i]],center)
r+=1
# riga sedute
put(ws,r,2,"Sedute",Font(bold=True))
for i in range(len(weeks)):
    put(ws,r,3+i,"2C+1P" if i<8 else "1C+gara",None,None,center)
r+=1
# gara
put(ws,r,2,"Eventi",Font(bold=True))
for i in range(len(weeks)):
    put(ws,r,3+i,"🏐 31/10" if i==8 else "",None,("FF6B6B" if i==8 else None),center)
r+=2

put(ws,r,2,"SETTIMANA TIPO (microciclo, esempio)", sub_font, bd=False); r+=1
head=["Giorno","Seduta","Contenuto"]
for i,h in enumerate(head): put(ws,r,2+i,h)
style_header(ws,r,len(head),start=2); r+=1
tipo=[
 ("Lunedì","Campo A — Atletica","Riscaldamento neuromuscolare · agilità/pliometria · condizionamento · prevenzione"),
 ("Martedì","(tecnica squadra)","—"),
 ("Mercoledì","Pesi — Forza principale","Fondamentali (squat/hinge/spinta/tirata) · core · prevenzione"),
 ("Giovedì","(tecnica squadra)","—"),
 ("Venerdì","Campo B — Forza campo + specifico","Riscaldamento neuromusc. · blocco forza (unilaterale/post./prev.) · gesto specifico · condizionamento specifico"),
 ("Sab/Dom","riposo / (gara in-season)","In fase gara: sab 31/10 1ª partita"),
]
for g,s,c in tipo:
    put(ws,r,2,g,Font(bold=True)); put(ws,r,3,s); ws.merge_cells(start_row=r,start_column=4,end_row=r,end_column=10); put(ws,r,4,c); r+=1
ws.freeze_panes="B4"

# ============================================================= LEGENDA & CARICHI
ws = wb.create_sheet("Legenda & Carichi")
widths(ws,{"A":4,"B":22,"C":50,"D":14,"E":12,"F":12})
put(ws,1,2,"LEGENDA & GESTIONE DEI CARICHI", title_font, bd=False)
r=3; put(ws,r,2,"Legenda",sub_font,bd=False); r+=1
leg=[
 ("RPE","Sforzo percepito 0–10 (10 = massimo, nessuna ripetizione di riserva). Guida principale del carico."),
 ("%1RM","Percentuale del massimale (per gli esercizi di sala). I kg sono calcolati su un 1RM di RIFERIMENTO."),
 ("S×R","Serie × Ripetizioni"),
 ("Rec","Recupero tra le serie"),
 ("Fasce","🔴 Rosso = regressioni (decondizionate/problemi motori) · 🟡 Giallo = base · 🟢 Verde = progressioni"),
 ("Riscaldamento neuromuscolare","Protocollo fisso tipo FIFA 11+ adattato: attivazione, controllo atterraggio, core (5–8')"),
]
for k,v in leg:
    put(ws,r,2,k,Font(bold=True)); ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=6); put(ws,r,3,v); r+=1
r+=1
put(ws,r,2,"1RM di riferimento (ASSUNZIONE — ricalcolare sul reale!)",sub_font,bd=False); r+=1
put(ws,r,2,"NB: valori medi ipotizzati per atleta Serie D. I kg nelle schede derivano da questi. Sostituire col 1RM (o stima sub-massimale) di ogni atleta.",Font(italic=True,color="B5540A")); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6); r+=1
for i,h in enumerate(["Esercizio","1RM rif. (kg)","60%","70%","80%"]): put(ws,r,2+i,h)
style_header(ws,r,5,start=2); r+=1
for ex,v in RM.items():
    put(ws,r,2,ex,Font(bold=True)); put(ws,r,3,v,None,None,center)
    put(ws,r,4,kg(60,v),None,None,center); put(ws,r,5,kg(70,v),None,None,center); put(ws,r,6,kg(80,v),None,None,center); r+=1
r+=1
put(ws,r,2,"Le 3 fasce — come si muovono",sub_font,bd=False); r+=1
fasce=[
 ("🔴 Rosso","Corpo libero/manubri leggeri, ROM controllato, atterraggi (non salti), supporti (TRX/box), volumi ridotti, RPE 5–6. Sale quando il movimento è pulito."),
 ("🟡 Giallo","Versione standard, RPE 6–8. Riferimento del programma."),
 ("🟢 Verde","Carico/bilanciere, salti e ri-salti, cluster, velocità, %1RM, RPE 7–9 (mai cedimento in preparazione)."),
]
for k,v in fasce:
    put(ws,r,2,k,Font(bold=True)); ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=6); put(ws,r,3,v); r+=1

# ============================================================= helper sheet fase
def gym_prog():
    # (esercizio, 1RM key or None) , dict week-> (sxr, pct, rpe)
    return [
     ("Back Squat","Squat",{1:("3×10",55,"6"),2:("4×8",68,"7"),3:("4×6",73,"7-8"),4:("3×6",65,"6"),5:("4×5",78,"8"),6:("5×4",82,"8"),7:("4×3",85,"8-9"),8:("3×3",80,"7 (velocità)")}),
     ("Panca manubri/bilanciere","Panca",{1:("3×10",55,"6"),2:("4×8",65,"7"),3:("4×6",72,"7-8"),4:("3×6",63,"6"),5:("4×5",75,"8"),6:("4×4",80,"8"),7:("4×3",83,"8-9"),8:("3×4",78,"7")}),
     ("RDL / Hip hinge","RDL",{1:("3×10",55,"6"),2:("3×8",62,"7"),3:("4×8",65,"7"),4:("2×8",58,"6"),5:("4×6",70,"7-8"),6:("4×6",73,"8"),7:("3×5",76,"8"),8:("3×5",70,"7")}),
     ("Trazioni/Lat o Row","",{1:("3×10",None,"7"),2:("4×8",None,"7-8"),3:("4×8",None,"8"),4:("2×8",None,"6"),5:("5×6",None,"8"),6:("5×6",None,"8"),7:("4×5",None,"8-9"),8:("4×5",None,"7")}),
    ]

def write_phase(name, weeks_idx, color, obj, extra_notes):
    ws = wb.create_sheet(name)
    widths(ws,{"A":3,"B":26,"C":12,"D":12,"E":16,"F":14,"G":30})
    put(ws,1,2,fase_label[color] if color in fase_label else name, title_font, bd=False)
    wk = ", ".join([f"S{i+1} ({fmt(weeks[i][0])}–{fmt(weeks[i][1])})" for i in weeks_idx])
    put(ws,2,2,f"Settimane: {wk}", Font(italic=True,color="666666"), bd=False)
    r=4
    put(ws,r,2,"Obiettivo della fase",sub_font,bd=False); r+=1
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7); put(ws,r,2,obj); r+=2

    # --- STRUTTURA SEDUTE (skeleton) ---
    put(ws,r,2,"STRUTTURA SEDUTE (skeleton · tempi indicativi su 60')",sub_font,bd=False); r+=1
    for titolo, blocchi in SKELETONS[color]:
        put(ws,r,2,titolo,Font(bold=True,color=NAVY),PH[color]); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7); r+=1
        for i,h in enumerate(["Blocco (min)","Esercizio","S×R","RPE","Carico/Attrezzo","Note / Fasce"]):
            put(ws,r,2+i,h)
        style_header(ws,r,6,start=2); r+=1
        for row in blocchi:
            for i,v in enumerate(row): put(ws,r,2+i,v, None, None, (center if i in (2,3) else wrap))
            r+=1
        r+=1

    # --- PROGRESSIONE SETTIMANALE SALA (kg+RPE) ---
    put(ws,r,2,"PROGRESSIONE SALA — carichi settimana per settimana (kg su 1RM rif. + RPE)",sub_font,bd=False); r+=1
    hdr=["Esercizio"]+[f"S{i+1}" for i in weeks_idx]
    for i,h in enumerate(hdr): put(ws,r,2+i,h)
    style_header(ws,r,len(hdr),start=2); r+=1
    for ex,rmkey,prog in gym_prog():
        put(ws,r,2,ex,Font(bold=True))
        for j,i in enumerate(weeks_idx):
            sxr,pct,rpe = prog[i+1]
            if pct and rmkey:
                cell=f"{sxr}\n{pct}% ≈ {kg(pct,RM[rmkey])}kg\nRPE {rpe}"
            else:
                cell=f"{sxr}\nRPE {rpe}"
            put(ws,r,3+j,cell,Font(size=9),PH[color],center)
        r+=1
    r+=1
    if extra_notes:
        put(ws,r,2,"Note di progressione (campo/pliometria/condizionamento)",sub_font,bd=False); r+=1
        for n in extra_notes:
            ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7); put(ws,r,2,"• "+n); r+=1
    ws.freeze_panes="C5"
    return ws

# ---- skeleton sedute per fase (Campo A, Pesi, Campo B) ----
def sk_common_prev():
    return ["8–10'","Prevenzione: cuffia (extrarot. 2×15/lato), Face pull 2×15, caviglia (calf 2×15 + balance 2×20s)","—","leggero","elastico","spalla+caviglia, in TUTTE le sedute"]

SKELETONS = {
 "Accumulo":[
  ("CAMPO A — Atletica (60')",[
    ["10'","Riscaldamento neuromuscolare (attivazione, mobilità, core, atterraggi)","—","4-5","corpo libero/skimmy","protocollo fisso (prev. LCA)"],
    ["12'","Motricità/andature + skip tra cinesini; intro agilità (slalom coni)","3×15m","tecnico","—","🔴 camminato · 🟢 a velocità"],
    ["12'","Pliometria: atterraggi soft landing → salti bassi","3×5","estensivo","box basso","🔴 solo atterraggi · 🟢 salto→arresto"],
    ["16'","Circuito forza-cond. 3 giri 40/20 (squat elastico, spinta TRX, affondo, row elastico, dead bug)","3× giro","6-7","elastico/TRX","🔴 2 giri · 🟢 4 giri"],
    ["7'","Condizionamento: navette 15m","6-8×","moderato","—","🔴 4× · 🟢 10×"],
    sk_common_prev(),
  ]),
  ("PESI — Forza principale (60')",[
    ["10'","Riscaldamento + avvicinamento sui fondamentali","2×8","tecnico","rampa","—"],
    ["35'","Fondamentali: Squat · RDL/Hinge · Panca · Trazioni/Row (vedi progressione sala)","vedi tab.","6-8","kg/RPE","🔴 goblet/box, no bilanciere · 🟢 %1RM"],
    ["10'","Core & prevenzione (dead bug 3×8/lato, pallof 3×10/lato)","3×","6","corpo libero","+ extrarot. spalla, calf"],
    ["5'","Cooldown mobilità","—","—","—","—"],
  ]),
  ("CAMPO B — Forza campo + specifico (60')",[
    ["10'","Riscaldamento neuromuscolare","—","4-5","corpo libero","protocollo fisso"],
    ["22'","BLOCCO FORZA CAMPO: Bulgarian/Step-up 3×8/gamba · Hip hinge/1-leg RDL 3×8/lato · Nordic ecc. 3×6 · Ponte 1 gamba 3×8 · Carry 3×30m","3×","7","manubri/elastici/box","2° stimolo forza settimanale · 🔴 corpo libero"],
    ["12'","Sport-specifico: gesto (battuta/attacco a secco) + spostamenti/arresto","tecnico","—","—","controllo decelerazione"],
    ["8'","Condizionamento generale (circuito/navette)","—","moderato","—","🟢 +volume"],
    sk_common_prev(),
  ]),
 ],
 "Deload":[
  ("CAMPO A — Scarico qualità (60')",[
    ["10'","Riscaldamento neuromuscolare","—","4","corpo libero","—"],
    ["15'","Motricità + atterraggi di QUALITÀ (no salti max)","2×5","tecnico","box basso","volume −40%"],
    ["15'","Circuito 2 giri (ridotto)","2× giro","6","elastico/TRX","—"],
    ["10'","Mobilità/attivazione","—","leggero","—","recupero attivo"],
    sk_common_prev(),
  ]),
  ("PESI — Scarico (60')",[
    ["10'","Riscaldamento","—","tecnico","—","—"],
    ["30'","Fondamentali a volume ridotto (vedi progressione S4)","vedi tab.","6","~65% / RPE6","niente cedimento"],
    ["10'","Core leggero + prevenzione","2×","5-6","corpo libero","—"],
    ["5'","Cooldown","—","—","—","—"],
  ]),
  ("CAMPO B — Scarico + mobilità (60')",[
    ["10'","Riscaldamento neuromuscolare","—","4","corpo libero","—"],
    ["18'","Forza campo ridotta (2 serie) + tecnica gesto","2×","6","manubri leggeri","—"],
    ["12'","Mobilità globale + core","—","leggero","—","recupero"],
    sk_common_prev(),
  ]),
 ],
 "Sviluppo":[
  ("CAMPO A — Atletica/Pliometria (60')",[
    ["10'","Riscaldamento neuromuscolare","—","5","corpo libero","protocollo fisso"],
    ["15'","Agilità a velocità (T-drill, cambi direzione su stimolo)","4-5×","media-alta","coni/skimmy","🔴 lento · 🟢 reattivo"],
    ["15'","Pliometria estensiva → clusterizzata (ostacoli, cluster 4+4+4, rec ampio)","vedi note","qualità","ostacoli/box","gate: forza base + tecnica"],
    ["12'","Condizionamento SPECIFICO: salti/spostamenti ripetuti (work:rest 1:2-1:3)","5-6×","alto","—","simulazione attacco/muro"],
    sk_common_prev(),
  ]),
  ("PESI — Forza (intensità) (60')",[
    ["10'","Riscaldamento + rampa","2×5","tecnico","rampa","—"],
    ["35'","Fondamentali (vedi progressione: intensità ↑, reps ↓)","vedi tab.","8-9","kg/RPE","🟢 %1RM · 🔴 RPE7 tecnica"],
    ["10'","Core/prevenzione (anti-rotazione, Nordic progressione)","3×","7","—","focus LCA"],
    ["5'","Cooldown","—","—","—","—"],
  ]),
  ("CAMPO B — Forza campo + trasferimento (60')",[
    ["10'","Riscaldamento neuromuscolare","—","5","corpo libero","—"],
    ["20'","Forza campo (carico/tempo ↑) + salto caricato leggero (🟢)","3-4×","7-8","manubri/box","trasferimento forza→potenza"],
    ["15'","Sport-specifico: gesto a intensità + situazioni brevi di gioco","—","specifico","—","—"],
    ["8'","Condizionamento specifico","5×","alto","—","🟢 +intensità"],
    sk_common_prev(),
  ]),
 ],
 "Realizzazione":[
  ("CAMPO A — Reattività/Qualità (60')",[
    ["10'","Riscaldamento neuromuscolare","—","5","corpo libero","—"],
    ["15'","Pliometria di QUALITÀ (CMJ max intent, poche rip., freschezza)","3×4","alta/qualità","box","taper: volume basso"],
    ["15'","Agilità reattiva + gesto specifico","—","alta","coni","—"],
    ["12'","Gioco/situazioni","—","specifico","—","—"],
    sk_common_prev(),
  ]),
  ("PESI — Mantenimento/velocità (60')",[
    ["10'","Riscaldamento + rampa","—","tecnico","—","—"],
    ["30'","Fondamentali a volume basso, focus VELOCITÀ (vedi S8)","3×3-5","7","~80% / RPE7","niente massimali · qualità"],
    ["10'","Core/prevenzione mantenimento","2×","6","—","—"],
    ["5'","Cooldown","—","—","—","—"],
  ]),
  ("CAMPO B — Rifinitura + TEST (60')",[
    ["10'","Riscaldamento neuromuscolare","—","5","corpo libero","—"],
    ["25'","RE-TEST: CMJ, spike/block reach, single-leg hop (simmetria), plank","—","test","—","vedi scheda Test & Marker"],
    ["15'","Gesto specifico leggero + mobilità","—","leggero","—","pronte alla gara"],
    ["10'","Prevenzione (mantenimento)","2×","leggero","elastico","—"],
  ]),
 ],
}

# scrivi le fasi
write_phase("F1 Accumulo (S1-3)",[0,1,2],"Accumulo",
  "Valutare e mettere in sicurezza (screening, atterraggi, mobilità), costruire capacità di lavoro e forza generale; introdurre agilità e salti bassi. Carichi bassi, qualità massima; volume crescente S1→S3.",
  ["Pliometria: S1 atterraggi/soft landing · S2 atterraggi + salti bassi 3×5 · S3 salti bassi + intro ostacolo. Gate per salire: atterraggio allineato e indolore.",
   "Condizionamento generale: navette 6→8; circuiti 3 giri. RPE 6→7.",
   "Forza campo (Campo B): volume 2→3-4 serie su unilaterale/catena post./Nordic. È il 2° stimolo forza settimanale.",
   "Prevenzione: volume in crescita; Nordic eccentrico progressione settimanale (focus LCA)."])

write_phase("Deload (S4)",[3],"Deload",
  "Scarico programmato (riposo pulsato): −40% volume, intensità mantenuta ma senza cedimento. Recupero, consolidamento tecnico e RE-CHECK delle fasce prima del blocco intenso.",
  ["Pliometria: solo atterraggi di qualità. Nessun salto massimale.",
   "Sala a volume ridotto (~65% / RPE6). Campo a 2 serie.",
   "Obiettivo: arrivare fresche e riassegnare Rosso/Giallo/Verde per la Fase 2."])

write_phase("F2 Sviluppo (S5-7)",[4,5,6],"Sviluppo",
  "Sviluppo della forza (intensità ↑, reps ↓), pliometria estensiva strutturata → clusterizzata, condizionamento SPECIFICO volley, avvio trasferimento a potenza. Volume medio, intensità alta.",
  ["Pliometria: S5 estensiva strutturata · S6 clusterizzata (4+4+4, rec ampio) · S7 reattività + (🟢) intro intensiva leggera (drop basso) SOLO con gate forza (squat ~2×BW o profilo F-V).",
   "Condizionamento SPECIFICO: salti/spostamenti ripetuti, simulazione attacco/muro, work:rest 1:2–1:3.",
   "Forza campo: carico/tempo ↑ + salto caricato leggero (trasferimento).",
   "Prevenzione: mantenimento su volume alto; Nordic ai livelli massimi del ciclo."])

write_phase("F3 Realizzazione (S8)",[7],"Realizzazione",
  "Taper: volume basso, qualità e velocità. Re-test e rifinitura. Obiettivo: massima freschezza e prontezza per la 1ª gara (31/10).",
  ["Sala: 3×3-5 focus velocità (~80%), niente massimali.",
   "Pliometria: solo qualità (CMJ max intent, poche rip.).",
   "Chiusura con RE-TEST completo (vedi scheda Test & Marker)."])

# ============================================================= SETTIMANA TIPO IN-SEASON
ws = wb.create_sheet("Settimana Tipo (gara)")
widths(ws,{"A":3,"B":14,"C":26,"D":50})
put(ws,1,2,"SETTIMANA TIPO IN-SEASON (1ª gara sab 31/10)", title_font, bd=False)
put(ws,2,2,"Microciclo settimanale con partita nel weekend: mantenere, non affaticare.", Font(italic=True,color="666666"), bd=False)
r=4
for i,h in enumerate(["Giorno","Seduta atletica","Contenuto"]): put(ws,r,2+i,h)
style_header(ws,r,3,start=2); r+=1
inseason=[
 ("Lun","Scarico post nulla / attivazione","Mobilità + attivazione leggera (se sett. senza gara precedente, richiamo forza)"),
 ("Mar","(tecnica)","—"),
 ("Mer","Forza mantenimento (breve)","Fondamentali 3×3-4 @~80% RPE7 (freschi) + core/prevenzione — 40'"),
 ("Gio","(tecnica)","Prevenzione integrata al riscaldamento tecnico"),
 ("Ven","Attivazione pre-gara","Neuromuscolare + qualche salto di qualità + mobilità (30-40')"),
 ("Sab 31/10","🏐 GARA","1ª partita"),
 ("Dom","Recupero","Scarico attivo / mobilità (opzionale)"),
]
for g,s,c in inseason:
    put(ws,r,2,g,Font(bold=True),("FF6B6B" if "GARA" in s else None)); put(ws,r,3,s); put(ws,r,4,c); r+=1

# ============================================================= TEST & MARKER
ws = wb.create_sheet("Test & Marker")
widths(ws,{"A":3,"B":20,"C":30,"D":26,"E":30})
put(ws,1,2,"TEST, MARKER E OBIETTIVI PER FASE", title_font, bd=False)
put(ws,2,2,"Criteri per considerare 'superata' ogni fase. Test semplici, ripetibili, sport-specifici e attenti alla simmetria dx/sx (prevenzione LCA).", Font(italic=True,color="666666"), bd=False)
r=4
for i,h in enumerate(["Fase","Obiettivi da raggiungere","Test / Marker","Criterio di superamento"]): put(ws,r,2+i,h)
style_header(ws,r,4,start=2); r+=1
tests=[
 ("Ingresso (S1)","Fotografare baseline e assegnare le fasce","Overhead squat · Single-leg balance · Single-leg hop (simmetria) · Landing test (valgo) · Mobilità spalla · Plank max","Completato su tutte; fasce assegnate 🔴🟡🟢","Accumulo"),
 ("FASE 1 Accumulo (fine S3)","Movimento pulito, tolleranza al carico, forza generale avviata","Ri-osservazione landing · RPE medio seduta · Plank · esecuzione fondamentali","Atterraggio allineato e indolore; RPE in calo a pari lavoro; tecnica fondamentali pulita","Accumulo"),
 ("Deload (fine S4)","Recupero completo; fasce riassegnate","Percezione recupero (wellness) · qualità movimento","Atlete fresche, nessun dolore; fasce aggiornate","Deload"),
 ("FASE 2 Sviluppo (fine S7)","Incremento forza; reattività; condizione specifica","CMJ · carichi sala (kg) · tolleranza condizionamento specifico · Nordic","+ altezza CMJ vs baseline; +carico fondamentali; regge il lavoro specifico senza cali tecnici","Sviluppo"),
 ("FASE 3 Realizzazione (fine S8)","Freschezza + picco di qualità; pronte alla gara","RE-TEST: CMJ · Spike/Block reach · Single-leg hop (simmetria dx/sx) · Plank","CMJ ≥ baseline (freschezza); simmetria hop entro ~10%; nessun dolore → IDONEE alla 1ª gara","Realizzazione"),
]
for fa,ob,te,cr,col in tests:
    put(ws,r,2,fa,Font(bold=True),PH[col]); put(ws,r,3,ob); put(ws,r,4,te); put(ws,r,5,cr); r+=1
r+=2
put(ws,r,2,"Marker di gestione continua (tutte le fasi)",sub_font,bd=False); r+=1
for m in [
 "RPE seduta (carico interno) — se sale a pari lavoro esterno = affaticamento.",
 "Wellness breve (sonno, dolori, energia) — soprattutto in un gruppo femminile; autoregolare il carico.",
 "Dolore/fastidio articolare (ginocchio, spalla, caviglia) — STOP e regressione se presente.",
 "Simmetria arti (hop test) — differenza >10-15% = rischio, individualizzare.",
]:
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5); put(ws,r,2,"• "+m); r+=1

# salva
out="/home/user/schede-allenamento/programmi/preseason_volley_serieD/Programma_PreSeason_VolleyD.xlsx"
import os; os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("Salvato:", out)
print("Fogli:", wb.sheetnames)
