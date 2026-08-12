# -*- coding: utf-8 -*-
import datetime as dt, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
NAVY="0B4A6F"; WHITE="FFFFFF"
def fillp(c): return PatternFill("solid", fgColor=c)
PH={"Accumulo":"C6E0B4","Deload":"FFE699","Sviluppo":"F8CBAD","Realizzazione":"BDD7EE"}
thin=Side(style="thin",color="B9C6CE"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
hf=Font(bold=True,color=WHITE,size=10); hfill=fillp(NAVY)
tf=Font(bold=True,color=NAVY,size=14); sf=Font(bold=True,color=NAVY,size=11)
wrap=Alignment(wrap_text=True,vertical="top"); center=Alignment(horizontal="center",vertical="center",wrap_text=True)
def put(ws,r,c,v,font=None,fc=None,al=wrap,bd=True):
    cell=ws.cell(row=r,column=c,value=v)
    if font:cell.font=font
    if fc:cell.fill=fillp(fc)
    if al:cell.alignment=al
    if bd:cell.border=border
    return cell
def hdr(ws,row,cols,start=2):
    for i,h in enumerate(cols): put(ws,row,start+i,h,hf,None,center)
def widths(ws,d):
    for k,v in d.items(): ws.column_dimensions[k].width=v

start=dt.date(2026,8,31)
weeks=[(start+dt.timedelta(days=7*i), start+dt.timedelta(days=7*i+6)) for i in range(8)]
fase_of=["Accumulo","Accumulo","Accumulo","Deload","Sviluppo","Sviluppo","Sviluppo","Realizzazione"]
fase_lab={"Accumulo":"FASE 1 · Accumulo","Deload":"Deload","Sviluppo":"FASE 2 · Sviluppo","Realizzazione":"FASE 3 · Realizzazione/Test"}
def fmt(d): return d.strftime("%d/%m")
RM={"Squat":55,"Panca":30,"RDL":65}
def kg(p,o): return round((p/100.0*o)/2.5)*2.5
def load(p,rm): return f"{p}% ≈ {kg(p,rm)}kg" if p else "—"

# ---------------- progressioni ----------------
SQ={1:("3×10",55,"6"),2:("4×8",68,"7"),3:("4×6",73,"7-8"),4:("3×6",65,"6"),5:("4×5",78,"8"),6:("5×4",82,"8"),7:("4×3",85,"8-9"),8:("3×3",80,"7 vel.")}
PA={1:("3×10",55,"6"),2:("4×8",65,"7"),3:("4×6",72,"7-8"),4:("3×6",63,"6"),5:("4×5",75,"8"),6:("4×4",80,"8"),7:("4×3",83,"8-9"),8:("3×4",78,"7")}
RD={1:("3×10",55,"6"),2:("3×8",62,"7"),3:("4×8",65,"7"),4:("2×8",58,"6"),5:("4×6",70,"7-8"),6:("4×6",73,"8"),7:("3×5",76,"8"),8:("3×5",70,"7")}
TI={1:("3×10","7"),2:("4×8","7-8"),3:("4×8","8"),4:("2×8","6"),5:("5×6","8"),6:("5×6","8"),7:("4×5","8-9"),8:("4×5","7")}
AF={1:("2×10/g","6"),2:("2×10/g","7"),3:("3×8/g","7"),4:("2×8/g","6"),5:("3×8/g","7"),6:("3×8/g","8"),7:("3×6/g","8"),8:("2×8/g","6")}
# campo A
AGI={1:"Andature + skip A/B tra cinesini",2:"Slalom coni + skip laterale",3:"Cambi direzione a L (decelerazione)",4:"Mobilità + skip (scarico)",5:"T-drill a velocità",6:"Cambi direzione su stimolo",7:"Reattività su stimolo + salto→arresto",8:"Agilità reattiva (qualità)"}
PLIO={1:("Atterraggi soft landing","3×5","estensivo"),2:("Atterraggi + salti bassi vert.","3×5","estensivo"),3:("Salti bassi + intro ostacolo","3×5","estensivo"),4:("Solo atterraggi di qualità","2×5","scarico"),5:("Estensiva strutturata (ostacoli)","4×5","qualità"),6:("Clusterizzata (4+4+4, rec ampio)","3 cluster","qualità"),7:("Reattività + (🟢) drop basso","4×4","alta qualità"),8:("CMJ max intent (freschezza)","3×4","alta/qualità")}
COND={1:("Navette 15m","6×","moderato"),2:("Navette 15m","8×","moderato"),3:("Navette + mini-circuito","8×","moderato"),4:("Condizionamento ridotto","4×","blando"),5:("SPECIFICO: salti/spost. ripetuti (1:2)","5×","alto"),6:("SPECIFICO (1:2)","6×","alto"),7:("SPECIFICO + gioco (1:3)","5×","alto"),8:("Minimale","3×","blando")}
# campo B forza campo
NORD={1:"3×4",2:"3×5",3:"3×6",4:"2×4",5:"3×6",6:"3×6 tempo",7:"3×5 +carico",8:"2×5"}
FCB_serie={1:"2-3",2:"3",3:"3-4",4:"2",5:"3-4",6:"4",7:"3-4",8:"2"}
GESTO={1:"Posizione + spostamenti base",2:"Gesto battuta a secco",3:"Gesto attacco a secco + rincorsa",4:"Tecnica leggera (scarico)",5:"Attacco/muro a secco a intensità",6:"Situazioni brevi di gioco",7:"Situazioni + reattività",8:"Gesto leggero (rifinitura pre-gara)"}

PREV=["8-10'","PREVENZIONE (fissa): cuffia extrarot. 2×15/lato · Face pull 2×15 · Y-T-W 2×8 · caviglia calf 2×15 + balance 2×20s","2-3×","leggero","elastico","1 serie","2×20"]

def campoA(w):
    f=fase_of[w-1]
    pl=PLIO[w]; cd=COND[w]
    rows=[
     ["10'","Riscaldamento neuromuscolare (attivazione, mobilità, controllo atterraggio, core)","1×","4-5","corpo libero/skimmy","= ","= "],
     ["12'",AGI[w],"3×15m","tecnico","coni/skimmy","camminato/lento","a velocità"],
     ["13'","Pliometria: "+pl[0],pl[1],pl[2],"box/ostacolo","solo atterraggi","salto→arresto / +ampiezza"],
    ]
    if w<=3:
        rows.append(["15'","Circuito forza-cond. 3 giri 40/20 (squat elastico · spinta TRX · affondo · row elastico · dead bug)","3× giro","6-7","elastico/TRX","2 giri","4 giri"])
    else:
        rows.append(["12'","Circuito forza-cond. (mantenimento)","2× giro","6-7","elastico/TRX","1-2 giri","3 giri"])
    rows.append(["7'","Condizionamento: "+cd[0],cd[1],cd[2],"—","meno volume","più volume/intensità"])
    rows.append(PREV)
    return rows

def pesi(w):
    sq=SQ[w];pa=PA[w];rd=RD[w];ti=TI[w];af=AF[w]
    dl=" (deload)" if w==4 else ""
    return [
     ["10'","Riscaldamento + rampa sui fondamentali","2×6","tecnico","a vuoto→target","solo corpo libero","rampa fino a target"],
     ["35'","Back Squat"+dl,sq[0],sq[2],load(sq[1],RM["Squat"]),"goblet/box squat RPE7","bilanciere · %1RM"],
     ["35'","RDL / Hip hinge",rd[0],rd[2],load(rd[1],RM["RDL"]),"manubri leggeri (tecnica)","bilanciere · +carico"],
     ["35'","Panca manubri/bilanciere",pa[0],pa[2],load(pa[1],RM["Panca"]),"push-up inclinati","bilanciere · +carico"],
     ["35'","Trazioni assistite / Lat / Row",ti[0],ti[1],"corpo libero/assist.","row seduto elastico","trazioni libere/zavorra"],
     ["—","Affondi manubri (accessorio)",af[0],af[1],"manubri","corpo libero/appoggio","+carico"],
     ["10'","Core: Dead bug 3×8/lato + Pallof 3×10/lato","3×","6","corpo libero","braccia sole","+ elastico / in affondo"],
     ["10'","Prevenzione: extrarot. spalla 2×15/lato + calf 2×15","2×","leggero","elastico","1×15","2×20"],
     ["5'","Cooldown mobilità","—","—","—","—","—"],
    ]

def campoB(w):
    ser=FCB_serie[w]; nord=NORD[w]; gs=GESTO[w]; cd=COND[w]
    dl=" — scarico (2 serie)" if w==4 else ""
    rows=[
     ["10'","Riscaldamento neuromuscolare","1×","4-5","corpo libero","= ","= "],
     ["6'","Forza campo: Bulgarian/Step-up"+dl,f"{ser}×8/gamba","7","manubri/box","corpo libero","+carico/tempo"],
     ["6'","Forza campo: Hip hinge / Single-leg RDL",f"{ser}×8/lato","7","manubri/elastico","bastone tecnica","+carico"],
     ["5'","Forza campo: Nordic eccentrico (LCA)",nord,"—","corpo libero","assistito/ROM ridotto","tempo/+ROM"],
     ["5'","Forza campo: Ponte 1 gamba + Carry",f"{ser}×8 + 30m","6-7","manubri","corpo libero","+carico"],
    ]
    if w>=5:
        rows.append(["3'","Salto caricato leggero (trasferimento forza→potenza)","3×4","qualità","manubri leggeri/box","— (no)","🟢 abilitato con gate forza"])
    rows.append(["12'","Sport-specifico: "+gs,"tecnico","—","—","semplificato","+ intensità/velocità"])
    rows.append(["8'","Condizionamento: "+cd[0],cd[1],cd[2],"—","meno volume","+ intensità"])
    rows.append(PREV)
    return rows

SESS=[("Campo A — Atletica","Lun",campoA),("Pesi — Forza principale","Mer",pesi),("Campo B — Forza campo + specifico","Ven",campoB)]
COLS=["Blocco (min)","Esercizio","S×R (🟡 base)","RPE","Carico (kg / %1RM)","🔴 Regressione","🟢 Progressione","Note & modifiche per atleta"]

# ---------------- INDICE ----------------
ws=wb.active; ws.title="Indice"
widths(ws,{"A":4,"B":10,"C":13,"D":13,"E":26,"F":10,"G":40})
put(ws,1,2,"SEDUTE ESPLOSE — Pre-season Volley Serie D (FIGLIO)",tf,bd=False)
put(ws,2,2,"File operativo · 24 sedute · companion del PADRE 'Programma_PreSeason_VolleyD.xlsx' · v0.2",Font(italic=True,color="666666"),bd=False)
r=4; hdr(ws,r,["#","Settimana","Fase","Giorno/Tipo","Foglio","Contenuto sintetico"]); r+=1
n=0
for w in range(1,9):
    f=fase_of[w-1]
    for titolo,giorno,_ in SESS:
        n+=1
        put(ws,r,2,n,None,None,center)
        put(ws,r,3,f"S{w} ({fmt(weeks[w-1][0])})",None,PH[f])
        put(ws,r,4,fase_lab[f],None,PH[f])
        put(ws,r,5,f"{giorno} · {titolo.split(' — ')[0]}")
        put(ws,r,6,f"S{w}",None,None,center)
        put(ws,r,7,titolo.split(' — ')[1] if ' — ' in titolo else titolo)
        r+=1
ws.freeze_panes="B5"

# ---------------- FOGLI SETTIMANA ----------------
for w in range(1,9):
    f=fase_of[w-1]
    ws=wb.create_sheet(f"S{w}")
    widths(ws,{"A":3,"B":10,"C":34,"D":12,"E":8,"F":16,"G":22,"H":24,"I":30})
    put(ws,1,2,f"SETTIMANA {w} · {fase_lab[f]}",tf,None,None,False)
    put(ws,2,2,f"{fmt(weeks[w-1][0])}–{fmt(weeks[w-1][1])} · 3 sedute (2 Campo + 1 Pesi) · ~60'",Font(italic=True,color="666666"),bd=False)
    r=4
    for si,(titolo,giorno,fn) in enumerate(SESS,1):
        num=(w-1)*3+si
        put(ws,r,2,f"SEDUTA {num} · {giorno} · {titolo}",Font(bold=True,color=NAVY,size=11),PH[f]);
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=9); r+=1
        hdr(ws,r,COLS); r+=1
        for row in fn(w):
            for i,val in enumerate(row):
                al=center if i in (2,3) else wrap
                put(ws,r,2+i,val,None,None,al)
            put(ws,r,2+7,"",None,"FFFDE7")  # colonna note (gialla chiara, vuota)
            r+=1
        r+=1
    ws.freeze_panes="B4"

# ---------------- REGISTRO MODIFICHE ----------------
ws=wb.create_sheet("Registro modifiche")
widths(ws,{"A":3,"B":12,"C":18,"D":10,"E":24,"F":30,"G":26})
put(ws,1,2,"REGISTRO MODIFICHE INDIVIDUALI",tf,bd=False)
put(ws,2,2,"Storico delle personalizzazioni per atleta (tracciabilità). Compilare in campo/palestra.",Font(italic=True,color="666666"),bd=False)
r=4; hdr(ws,r,["Data","Atleta","Seduta","Esercizio","Modifica effettuata","Motivo"]); r+=1
for _ in range(40):
    for c in range(2,8): put(ws,r,c,"")
    r+=1
ws.freeze_panes="B5"

out="/home/user/schede-allenamento/programmi/preseason_volley_serieD/Sedute_PreSeason_VolleyD.xlsx"
wb.save(out)
print("Salvato:",out); print("Fogli:",wb.sheetnames)
