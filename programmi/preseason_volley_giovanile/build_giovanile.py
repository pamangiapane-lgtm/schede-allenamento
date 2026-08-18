#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Genera Programma_PreSeason_VolleyGiovanile.xlsx"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/user/schede-allenamento/programmi/preseason_volley_giovanile/Programma_PreSeason_VolleyGiovanile.xlsx"

DARK = "1F4E79"; MID = "2E75B6"; GREY = "808080"; REDF = "C00000"
U14C = "C6E0B4"; U17C = "F8CBAD"
WHITE = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

def fill(hexv): return PatternFill("solid", fgColor=hexv)

wb = openpyxl.Workbook()

def section_title(ws, cell, text, size=14):
    c = ws[cell]; c.value = text; c.font = Font(size=size, bold=True, color=DARK)

def hdr_row(ws, row, headers, ncols=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h); cell.fill = fill(MID); cell.font = WHITE; cell.alignment = CTR; cell.border = BORD

def note_block(ws, start_row, start_col, end_col, lines, title=None, title_color=MID):
    r = start_row
    if title:
        c = ws.cell(r, start_col, title); c.fill = fill(title_color); c.font = WHITE; c.alignment = LFT
        ws.merge_cells(start_row=r, start_column=start_col, end_row=r, end_column=end_col)
        r += 1
    for ln in lines:
        cc = ws.cell(r, start_col, "•  " + ln); cc.alignment = LFT
        ws.merge_cells(start_row=r, start_column=start_col, end_row=r, end_column=end_col)
        r += 1
    return r

# ============================================================ PANORAMICA
ws = wb.active; ws.title = "Panoramica"
ws.sheet_view.showGridLines = False
ws["A1"] = "PRE-SEASON VOLLEY FEMMINILE GIOVANILE — U14 e U17"
ws["A1"].font = Font(size=16, bold=True, color=DARK)
ws["A2"] = "3 settimane · 3 sedute/sett · 30–45′ · guida per i tecnici, esecuzione in autonomia · v0.3"
ws["A2"].font = Font(size=10, italic=True, color=GREY)
ws.merge_cells("A1:H1"); ws.merge_cells("A2:H2")

ws["A4"] = "⚠️  Atlete SENZA esperienza di preparazione atletica. Obiettivo: muoversi bene e in sicurezza, non la prestazione."
ws["A4"].font = Font(bold=True, color=REDF); ws.merge_cells("A4:H4")
ws["A5"] = "Nessun infortunio/limitazione noto (da riconfermare). Dolore articolare = stop su quell'esercizio."
ws["A5"].font = Font(italic=True, color=REDF); ws.merge_cells("A5:H5")

r = 7
section_title(ws, f"A{r}", "CALENDARIO E SUPERFICI"); r += 1
hdr_row(ws, r, ["Settimana", "Superficie", "Attrezzatura", "Tema"]); r += 1
cal = [
    ("W1", "Spiaggia (sabbia, a piedi scalzi)", "corpo libero, palloni, cinesini/linee", "Fondamenta: imparare i movimenti, atterraggi morbidi, equilibrio"),
    ("W2", "Spiazzale / strada", "+ muretto (salite, appoggio, salti bassi)", "Costruzione: consolidare, introdurre il muretto, più agilità"),
    ("W3", "Spiazzale / strada", "+ muretto", "Espressione: più reattività mantenendo la qualità, mini-test finale"),
]
for i, row in enumerate(cal):
    rr = r + i
    for c, val in enumerate(row, 1):
        cell = ws.cell(rr, c, val); cell.border = BORD; cell.alignment = LFT if c > 1 else CTR
    ws.cell(rr, 1).fill = fill([U14C, "FFE699", U17C][i] if False else ["BDD7EE","C6E0B4","F8CBAD"][i])
r += len(cal) + 2

section_title(ws, f"A{r}", "STRUTTURA FISSA DELLA SEDUTA (~30–45′)"); r += 1
hdr_row(ws, r, ["Fase", "Durata", "Contenuto"]); r += 1
struct = [
    ("A. Riscaldamento (fisso)", "8–10′", "mobilità + attivazione + andature — vedi scheda 'Riscaldamento & Sicurezza'"),
    ("B. Parte centrale", "18–28′", "il blocco della giornata (Forza / Salti / Condizionamento) — dosi per settimana nelle schede sedute"),
    ("C. Prevenzione + defaticamento", "5–7′", "2–3 esercizi di prevenzione + mobilità/respirazione — vedi scheda 'Prevenzione'"),
]
for i, row in enumerate(struct):
    rr = r + i
    for c, val in enumerate(row, 1):
        cell = ws.cell(rr, c, val); cell.border = BORD; cell.alignment = LFT if c > 1 else CTR
r += len(struct) + 1
ws.cell(r, 1, "Nota: ogni seduta riporta un tempo indicativo totale. Se si sfora: taglia prima il volume, poi le stazioni — mai riscaldamento/prevenzione.").font = Font(italic=True, size=9, color=GREY)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r += 3

section_title(ws, f"A{r}", "LE 6 SEDUTE (3 per gruppo)"); r += 1
hdr_row(ws, r, ["Gruppo", "Seduta", "Focus", "RPE target"]); r += 1
sess_idx = [
    ("U14", "S1", "Forza & Controllo", "4–6"),
    ("U14", "S2", "Salti & Agilità", "4–6"),
    ("U14", "S3", "Condizionamento & Gioco", "4–6"),
    ("U17", "S1", "Forza & Controllo", "5–7"),
    ("U17", "S2", "Salti & Agilità", "5–7"),
    ("U17", "S3", "Condizionamento & Gioco", "5–7"),
]
for i, row in enumerate(sess_idx):
    rr = r + i
    for c, val in enumerate(row, 1):
        cell = ws.cell(rr, c, val); cell.border = BORD; cell.alignment = LFT if c == 3 else CTR
    ws.cell(rr, 1).fill = fill(U14C if row[0] == "U14" else U17C)
r += len(sess_idx) + 2

section_title(ws, f"A{r}", "VISIONE PLURIENNALE — QUESTO BLOCCO È LA TAPPA ①"); r += 1
hdr_row(ws, r, ["Tappa", "Quando", "Chi", "Temi centrali"]); r += 1
tappe = [
    ("① Alfabetizzazione motoria", "questo blocco — 3 sett.", "U14 + U17", "pattern fondamentali, atterraggi sicuri, fiducia nel movimento — SEI QUI"),
    ("② Multilateralità e consolidamento", "autunno-inverno, ~8–10 sett.", "U14 + U17 (differenziato)", "più schemi motori, pliometria estensiva clusterizzata, primi carichi leggeri"),
    ("③ Sviluppo forza e trasferimento", "primavera / pre-season anno 2", "U14 + U17 (percorsi divergenti)", "forza progressiva, pliometria intensiva con gate (U17), niente intensiva pre-pubere (U14)"),
    ("④ Specializzazione e integrazione", "anno 2+", "U17 (U14 in transizione)", "periodizzazione in-season, gestione carico gara, prevenzione cronica"),
]
for i, row in enumerate(tappe):
    rr = r + i
    for c, val in enumerate(row, 1):
        cell = ws.cell(rr, c, val); cell.border = BORD; cell.alignment = LFT if c in (1,4) else CTR
r += len(tappe) + 1
ws.cell(r, 1, "Dettaglio dei 5 mesocicli dell'Anno 1 (obiettivi misurabili) nella scheda 'Mesocicli Anno 1'.").font = Font(italic=True, size=9, color=GREY)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

for col, w in {"A": 30, "B": 24, "C": 34, "D": 46}.items():
    ws.column_dimensions[col].width = w

# ============================================================ MESOCICLI ANNO 1
mz = wb.create_sheet("Mesocicli Anno 1")
mz.sheet_view.showGridLines = False
section_title(mz, "A1", "IL PRIMO ANNO, MESE PER MESE — 5 MESOCICLI CON OBIETTIVI MISURABILI")
mz["A2"] = "M1 = Tappa ①; M2+M3 = Tappa ②; M4+M5 = inizio Tappa ③ (si completa nella pre-season Anno 2)"
mz["A2"].font = Font(italic=True, size=10, color=GREY)
mz.merge_cells("A1:G1"); mz.merge_cells("A2:G2")

r = 4
hdr_row(mz, r, ["Mesociclo", "Periodo", "Cosa miglioriamo / introduciamo", "Obiettivi U14", "Obiettivi U17", "Strada metodologica", "Serve per"])
r += 1
mesocicli = [
    ("M1 — Pre-season: Alfabetizzazione motoria\n(QUESTO BLOCCO)", "Settembre · 3 sett.",
     "Nessun precedente: primo contatto assoluto. Si parte da zero, si costruiscono i prerequisiti.",
     "Tecnica nuovo movimento: squat a corpo libero (profondità, ginocchia allineate, no dolore) · capacità di atterraggio (6/6 atterraggi bipodalici morbidi in sabbia) · equilibrio (20″ mono dx/sx senza scompensi)",
     "Forza (3×12 squat a corpo libero, tecnica pulita fino all'ultima serie) · tecnica nuovo movimento: hip hinge (schema corretto, rachide neutro) · capacità di atterraggio (6/6 bipodalici + hop&stick assistiti)",
     "Alfabetizzazione motoria pura: corpo libero, bassi volumi, sicurezza prima di tutto. Base di ogni LTAD.",
     "I pattern imparati qui (squat, hinge, atterraggio) sono il vocabolario su cui M2 costruirà variabilità."),
    ("M2 — Consolidamento e prima variabilità", "Ottobre–Novembre · ~8 sett.",
     "In M1 movimento in condizioni facilitate. Qui: qualità sotto variabilità (superficie dura, meno appoggio) + primi carichi leggerissimi.",
     "Miglioramento forza (push-up: 5/8 piegamenti mani più basse) · capacità di atterraggio (8 atterraggi bipodalici tecnicamente buoni su superficie dura) · tecnica nuovo movimento: atterraggio monopodalico libero (hop&stick senza appoggio)",
     "Forza (avvicinarsi a 4 serie da 8 squat con manubri leggerissimi, senza eccessiva fatica) · capacità di atterraggio (hop&stick libero, 4/5 per lato puliti) · condizionamento (reggere 15′ di circuito giocato)",
     "Una variabile alla volta (superficie, poi assistenza, poi carico) — mai due insieme.",
     "Reggere la tecnica sotto variabilità permette in M3 la pliometria strutturata senza rischio."),
    ("M3 — Multilateralità e pliometria in cluster", "Dicembre–Gennaio · ~8 sett.",
     "M2 ha consolidato sotto variabilità moderata. Qui: repertorio motorio più ampio + pliometria estensiva in cluster (metodo ELAV/Squillante).",
     "Tecnica nuovo movimento: lancio/ricezione con rotazione del tronco · capacità di atterraggio (pliometria in cluster 3×4, atterraggio pulito) · forza (2×10 affondi con manubri leggerissimi, ginocchio allineato)",
     "Forza (4×8 squat carico leggero-moderato, RPE ≤6) · capacità di atterraggio (pliometria in cluster 4×6, qualità stabile) · condizionamento sport-specifico (6-8 ripetute salto/spostamento, work:rest 1:2)",
     "Multilateralità + pliometria a tappe con gate di forza: non si passa avanti se la qualità cala.",
     "Repertorio motorio ampio = atleta adattabile. Un corpo che sa muoversi in tanti modi si infortuna meno."),
    ("M4 — Sviluppo della forza generale", "Febbraio–Marzo · ~8 sett.",
     "I percorsi U14/U17 iniziano a divergere più chiaramente. La forza passa da 'leggera' a 'generale strutturata'.",
     "Forza (2-3×10 squat con carico leggero reale, es. manubri 2-4 kg) · tecnica nuovo movimento: affondo con cambio di ritmo (agility) · capacità di atterraggio (mono da un piccolo balzo, non solo da fermo)",
     "Forza (completare 4 serie da 8 squat con carico moderato senza eccessiva fatica) · tecnica nuovo movimento: hip hinge caricato (RDL manubri) · pliometria (primi salti dal muretto, solo chi supera il gate di forza)",
     "Finestra di forza LTAD piena per U17, progressiva per U14 — il carico entra solo dove i criteri precedenti sono superati.",
     "Base di forza su cui in M5/Tappa③ si costruirà la conversione forza→potenza."),
    ("M5 — Transfer e ponte verso l'Anno 2", "Aprile–Maggio · ~6-8 sett.",
     "Integriamo tutto (pattern, pliometria, forza) in compiti più simili al gesto di gara. Prepariamo la pre-season Anno 2.",
     "Forza (3×10 squat carico leggero, tecnica mantenuta a fine seduta) · condizionamento (45′ allenamento giocato, intensità moderata) · re-test vs M1 su tutti i marker",
     "Forza (consolidare stabilmente 4×8 squat carico moderato, RPE ≤6) · pliometria (salti dal muretto puliti, volume in leggero aumento) · re-test vs M1 (CMJ/simmetria hop/qualità pattern)",
     "Chiusura per criteri, non solo calendario: si decide chi è pronta per il salto di livello.",
     "Il punto di arrivo di quest'anno è il punto di partenza dell'Anno 2 — meno tempo a insegnare le basi."),
]
MZCOL = ["BDD7EE", "C6E0B4", "FFE699", "F8CBAD", "D9D2E9"]
for i, row in enumerate(mesocicli):
    rr = r + i
    for c, val in enumerate(row, 1):
        cell = mz.cell(rr, c, val); cell.border = BORD; cell.alignment = LFT
    mz.cell(rr, 1).fill = fill(MZCOL[i]); mz.cell(rr, 1).font = Font(bold=True)
    mz.row_dimensions[rr].height = 110

for col, w in {"A": 22, "B": 16, "C": 30, "D": 40, "E": 40, "F": 30, "G": 30}.items():
    mz.column_dimensions[col].width = w

# ============================================================ RISCALDAMENTO & SICUREZZA
rs = wb.create_sheet("Riscaldamento & Sicurezza")
rs.sheet_view.showGridLines = False
section_title(rs, "A1", "RISCALDAMENTO FISSO (uguale per U14 e U17)")
rs["A2"] = "Si fa prima di ogni seduta. In sabbia i movimenti sono più lenti/faticosi: va bene, riduci le distanze."
rs["A2"].font = Font(italic=True, size=10, color=GREY)
rs.merge_cells("A1:D1"); rs.merge_cells("A2:D2")

r = 4
hdr_row(rs, r, ["Blocco", "Contenuto", "W1", "W2–W3"]); r += 1
warm = [
    ("1. Corsa/cammino attivo", "corsa leggera (in sabbia: cammino veloce)", "2–3′", "2–3′"),
    ("2. Mobilità dinamica (1 giro)", "circonduzioni caviglie/spalle · affondo+rotazione · apri-libro · slanci gamba", "1 giro", "1 giro"),
    ("3. Andature (10–15 m)", "skip basso · calciata dietro · passo laterale · corsa in progressione leggera", "1 giro (si spiega la tecnica)", "2 giri (più veloce, il tempo c'è)"),
    ("4. Attivazione", "8 mini-squat lenti · 8 ponti glutei · plank 15–20″", "come da tabella", "come da tabella"),
]
for i, row in enumerate(warm):
    rr = r + i
    for c, val in enumerate(row, 1):
        cell = rs.cell(rr, c, val); cell.border = BORD; cell.alignment = LFT if c <= 2 else CTR
r += len(warm) + 2

section_title(rs, f"A{r}", "SCALA RPE"); r += 1
r = note_block(rs, r, 1, 4, [
    "Scala 1–10 (1 = passeggiata, 10 = massimo).",
    "U14: RPE 4–6. U17: RPE 5–7. Devono finire stanchi ma non distrutti, e sempre in grado di parlare.",
], title=None)
r += 1

section_title(rs, f"A{r}", "REGOLE DI SICUREZZA VALIDE PER TUTTI"); r += 1
r = note_block(rs, r, 1, 4, [
    "Atterraggio morbido e silenzioso, ginocchia in linea con le punte dei piedi (mai a X). Priorità assoluta per il ginocchio femminile.",
    "Idratazione e pause: con caldo/sole accorcia il condizionamento e aumenta le pause.",
    "Dolore articolare (ginocchio, caviglia, spalla, schiena) = stop su quell'esercizio. Fastidio muscolare normale, dolore 'a punta' no.",
    "Progredire solo se la tecnica regge. Meglio una settimana in più sul facile che un infortunio.",
    "Riferire al preparatore chi ha dolori ricorrenti, forte asimmetria dx/sx, o non regge nemmeno le versioni facili.",
])
r += 1

section_title(rs, f"A{r}", "NOTE SULLE SUPERFICI"); r += 1
r = note_block(rs, r, 1, 4, [
    "Sabbia W1 (scalzi): ottima per atterraggi/equilibrio ma carica polpaccio/achilleo → volume salti basso, calf raise/mobilità caviglia in prevenzione, corsa breve. Scarpe se fastidio.",
    "Spiazzale/strada W2–W3 (dura): attenzione al volume dei salti. Muretto per salire/scendere controllati (non buttarsi giù). Erba se disponibile.",
])
for col, w in {"A": 26, "B": 46, "C": 20, "D": 26}.items():
    rs.column_dimensions[col].width = w

# ============================================================ SESSION SHEETS
COLS_SESS = ["Esercizio", "W1", "W2", "W3", "Guarda che…", "🔴 Più facile", "🟢 Più difficile"]
CW_SESS = [30, 20, 20, 22, 34, 26, 26]

def add_session(name, group, focus, rpe, rows, notes, tempo, color):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, f"{group} — {name.split(' ',1)[1] if ' ' in name else name} · {focus}").font = Font(size=13, bold=True, color=DARK)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS_SESS))
    ws.cell(2, 1, f"Durata 30–45′ · RPE target {rpe} · Tempo indicativo parte centrale: {tempo}").font = Font(italic=True, color=GREY)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS_SESS))
    tag = ws.cell(3, 1, group); tag.fill = fill(color); tag.font = Font(bold=True); tag.alignment = LFT
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(COLS_SESS))
    hr = 5
    hdr_row(ws, hr, COLS_SESS)
    cur = hr + 1
    for row in rows:
        for c, val in enumerate(row, 1):
            cell = ws.cell(cur, c, val); cell.border = BORD
            cell.alignment = LFT if c in (1, 5, 6, 7) else CTR
        cur += 1
    cur += 1
    for note_title, lines in notes:
        cur = note_block(ws, cur, 1, len(COLS_SESS), lines, title=note_title, title_color=color)
        cur += 1
    for c, w in enumerate(CW_SESS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A6"
    return ws

# --- U14 S1 ---
add_session("U14 S1", "U14", "Forza & Controllo", "4–6", "~16–18′", "",  # placeholder overwritten below
            [], []) if False else None

u14_s1_rows = [
    ("Squat a corpo libero", "2×8 lento (assistita)", "2×8 lento (autonoma)", "2×10 o 🟢 tempo 3″ discesa",
     "sedersi indietro, ginocchia in linea, schiena lunga", "squat su panca/muretto (tocca e risali)", "tempo 3″ in discesa"),
    ("Affondo camminato", "2×8/gamba (con appoggio)", "2×8/gamba", "2×10/gamba o 🟢 passo più lungo",
     "busto dritto, ginocchio non cade all'interno", "affondo sul posto tenendo appoggio", "passo più lungo"),
    ("Ponte glutei", "2×10", "2×10", "2×12 o 🟢 mono (fine W)",
     "spingi coi talloni, stringi i glutei", "—", "ponte a una gamba (fine W)"),
    ("Push-up inclinato (mani su rialzo)", "2×6", "2×7", "2×8",
     "corpo dritto, scendi controllato", "mani più in alto", "mani più in basso"),
    ("Plank", "2×15″", "2×20″", "2×20–25″",
     "corpo in linea, non a pancia giù", "plank sulle ginocchia", "+5–10″"),
    ("Equilibrio su una gamba*", "2×15″/gamba", "2×20″/gamba", "2×20–25″/gamba",
     "sguardo avanti, piede saldo", "occhi aperti, appoggio vicino", "occhi chiusi / superficie morbida"),
]
add_session(
    "U14 S1", "U14 — Forza & Controllo", "Forza & Controllo", "4–6", u14_s1_rows,
    [
        ("Progressione", ["* Equilibrio: fallo DURANTE il recupero degli altri esercizi — non aggiunge tempo.",
                           "Volume basso, niente carichi, niente salti d'urto. Tecnica prima di tutto."]),
        ("Stop se", ["dolore al ginocchio nello squat/affondo → riduci la profondità o passa al facile."]),
    ], "~16–18′", U14C)

u14_s2_rows = [
    ("Atterraggio morbido (saltello piccolo → congela 2″)", "2×6 (sabbia)", "2×6 (strada, cura tecnica)", "2×6–8",
     "atterraggio silenzioso, ginocchia in linea, ammortizza", "salta pochissimo", "salto un po' più alto (mai giù dal muretto)"),
    ("Hop & stick (saltello avanti mono, tieni l'atterraggio)", "2×4/gamba", "2×4/gamba", "2×4–6/gamba",
     "resta fermo 2″ senza sbilanciarti", "salto cortissimo", "distanza leggermente maggiore"),
    ("Saltelli sopra la linea (avanti-indietro/laterali)", "2×15″", "2×15″", "2×20″",
     "piedi veloci e leggeri, resta basso", "saltelli piccoli", "ritmo più rapido"),
    ("Skip + cambio direzione (a coppie, 'specchio')", "3×20″", "3×20″", "3×20–25″",
     "reagisci al compagno, appoggi puliti", "movimenti lenti", "reazioni più rapide"),
    ("Andatura laterale + stop", "2× andata/ritorno", "2× andata/ritorno", "3× andata/ritorno",
     "fermati controllata, non trascinare i piedi", "distanze corte", "fermati su una gamba"),
]
add_session(
    "U14 S2", "U14 — Salti & Agilità", "Salti & Agilità", "4–6", u14_s2_rows,
    [
        ("Note superficie", ["Sabbia W1: perfetta per gli atterraggi (attutisce).",
                              "Strada W2–3: riduci il numero di salti, cerca l'erba se possibile. Niente salti giù dal muretto per l'U14 (solo salite/appoggi)."]),
    ], "~14–16′", U14C)

u14_s3_rows_hdr = ["Blocco", "W1", "W2", "W3", "Note", "", ""]
u14_s3_rows = [
    ("Staffette / giochi con palla", "8′", "9′", "10′",
     "corsa, cambi direzione, divertimento; work:rest 1:2 (tanto recupero)", "", ""),
    ("Mini-circuito a stazioni (squat, saltelli, equilibrio, addome)", "1 giro", "1–2 giri", "2 giri (30″/30″)",
     "qualità, non gara di velocità", "", ""),
    ("Gioco libero con palla / partitella tecnica", "5′", "5′", "5′ (estendibile se avanza tempo)",
     "adesione e piacere del movimento", "", ""),
]
def add_session_freeform(name, group, focus, rpe, header, rows, notes, tempo, color):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, f"{group}").font = Font(size=13, bold=True, color=DARK)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
    ws.cell(2, 1, f"Durata 30–45′ · RPE target {rpe} · Tempo indicativo parte centrale: {tempo}").font = Font(italic=True, color=GREY)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(header))
    tag = ws.cell(3, 1, focus); tag.fill = fill(color); tag.font = Font(bold=True); tag.alignment = LFT
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(header))
    hr = 5
    hdr_row(ws, hr, header)
    cur = hr + 1
    for row in rows:
        for c, val in enumerate(row, 1):
            if val == "" and c > 5:
                continue
            cell = ws.cell(cur, c, val); cell.border = BORD
            cell.alignment = LFT if c in (1, 5) else CTR
        cur += 1
    cur += 1
    for note_title, lines in notes:
        cur = note_block(ws, cur, 1, len(header), lines, title=note_title, title_color=color)
        cur += 1
    widths = [34, 16, 16, 24, 40, 5, 5]
    for c, w in enumerate(widths[:len(header)], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return ws

add_session_freeform(
    "U14 S3", "U14 — Condizionamento & Gioco", "Condizionamento & Gioco", "4–6",
    ["Blocco", "W1", "W2", "W3", "Note"],
    [(r[0], r[1], r[2], r[3], r[4]) for r in u14_s3_rows],
    [
        ("Se il tempo stringe", ["taglia in quest'ordine: giri del circuito → durata staffette.",
                                  "NON tagliare il gioco libero: è il motivo per cui torneranno la prossima volta."]),
        ("Guarda che", ["con caldo/sabbia si stancano prima → allunga le pause. Obiettivo: finire contenti, non esausti."]),
    ], "~18–20′", U14C)

# --- U17 ---
u17_s1_rows = [
    ("Squat a corpo libero", "2×12", "2–3×12", "3×12 (o 🟢 tempo 3″/pausa in basso)",
     "profondità con schiena lunga, ginocchia in linea", "squat a panca/muretto", "tempo 3″ discesa o pausa in basso"),
    ("Affondo camminato", "2×10/gamba", "2–3×10/gamba", "3×10/gamba",
     "ginocchio stabile, non cade dentro", "affondo sul posto", "affondo in camminata + rotazione"),
    ("Hip hinge (stacco a corpo libero)", "2×10", "2–3×10", "3×10",
     "schiena dritta, bacino indietro, senti dietro la coscia", "mani scivolano sulle cosce come guida", "una gamba leggermente sollevata (fine W)"),
    ("Push-up (al muretto W2+ → a terra)", "2×8 (mani alte)", "2–3×8–10 (al muretto)", "3×8–10 (a terra se pronta)",
     "corpo dritto, gomiti ~45°", "mani su rialzo alto", "mani a terra / ginocchia"),
    ("Step-up sul muretto (dalla W2)", "— (niente muretto in spiaggia)", "2–3×8/gamba", "3×8/gamba",
     "sali spingendo di gamba, controlla la discesa", "gradino più basso", "salita più esplosiva (senza saltare giù)"),
    ("Plank + side plank", "2×20″+2×15″/lato", "2×25″+2×20″/lato", "2×30″+2×20″/lato",
     "bacino fermo, corpo in linea", "sulle ginocchia", "+10″ o solleva una gamba"),
]
add_session(
    "U17 S1", "U17 — Forza & Controllo", "Forza & Controllo", "5–7", u17_s1_rows,
    [
        ("Volume progressivo", ["W1 → 2 serie (si impara la tecnica) · W2 → 2–3 serie · W3 → 3 serie come da dosi."]),
        ("Consiglio organizzativo", ["Abbina a coppie in alternanza (superset): Squat+Push-up, Hip hinge+Plank — usa bene i recuperi, la seduta non si allunga."]),
    ], "~22–24′ con abbinamenti (fino a ~28′ in sequenza)", U17C)

u17_s2_rows = [
    ("Atterraggio morbido (salto verticale → congela 2″)", "3×6", "3×6", "3×6–8",
     "ammortizza, ginocchia in linea, silenzioso", "salto più basso", "salto un po' più alto"),
    ("Hop & stick monopodalico", "3×5/gamba", "3×5/gamba", "3×5–6/gamba",
     "fermo 2″, controllo caviglia/ginocchio", "salti corti", "distanza maggiore / laterale"),
    ("Balzi in avanti brevi (estensivo)", "3×5", "3×5", "3×5–6",
     "atterraggio pulito ad ogni balzo, non 'rimbalzo selvaggio'", "3 balzi", "5 balzi con più ampiezza"),
    ("Step-up esplosivo / salita muretto (W2); W3: salto giù basso + atterraggio morbido", "— (niente muretto in spiaggia)", "3×6 (salita esplosiva)", "3×6 (salto basso, solo se atterraggio perfetto)",
     "W3: muretto basso, atterra morbido, ginocchia in linea", "resta su salita controllata", "muretto leggermente più alto (solo se pronta)"),
    ("Agilità: navetta 5-10-5 / cambi direzione", "3×15–20″", "4×15–20″", "4×15–20″",
     "frena e riparti controllata, appoggi puliti", "distanze corte", "reazione a segnale del tecnico"),
]
add_session(
    "U17 S2", "U17 — Salti & Agilità", "Salti & Agilità", "5–7", u17_s2_rows,
    [
        ("Progressione salti (a tappe)", ["W1 solo atterraggi/soft landing in sabbia → W2 estensivo su strada + salite muretto → W3 estensivo + reattività leggera + primi salti bassi (solo chi atterra bene).",
                                            "Niente salti d'urto/alti: non è l'età né il livello."]),
    ], "~20–23′ (recuperi ampi voluti — qualità prima di tutto)", U17C)

u17_s3_rows = [
    ("Intervalli giocati / navette", "10′", "11′", "12′",
     "20–30″ lavoro / 30–45″ pausa × 6–8; oppure fartlek giocato con palla", "", ""),
    ("Circuito a stazioni (squat, step-up, saltelli, core, equilibrio)", "2 giri", "2 giri", "2 giri (40″/20″)",
     "mantieni la tecnica anche a fine giro", "", ""),
    ("Prevenzione estesa + gioco", "5′", "5′", "5′ (estendibile)",
     "vedi scheda Prevenzione + partitella", "", ""),
]
add_session_freeform(
    "U17 S3", "U17 — Condizionamento & Gioco", "Condizionamento & Gioco", "5–7",
    ["Blocco", "W1", "W2", "W3", "Note"],
    [(r[0], r[1], r[2], r[3], r[4]) for r in u17_s3_rows],
    [
        ("Se il tempo stringe", ["taglia in quest'ordine: giri del circuito (da 2 a 1) → durata intervalli.",
                                  "NON tagliare la prevenzione: qui si allena spalla e ginocchio."]),
        ("Guarda che", ["in W3 puoi alzare leggermente il ritmo, ma se la tecnica peggiora si torna indietro."]),
    ], "~20–22′", U17C)

# ============================================================ PREVENZIONE
pv = wb.create_sheet("Prevenzione")
pv.sheet_view.showGridLines = False
section_title(pv, "A1", "PREVENZIONE (focus femminile) — in ogni seduta, 5–7′")
pv["A2"] = "Rischio maggiore al ginocchio (LCA): controllo dell'atterraggio e forza di anca/coscia posteriore sono la miglior protezione. Scegli 2–3 esercizi a rotazione."
pv["A2"].font = Font(italic=True, size=10, color=GREY)
pv.merge_cells("A1:D1"); pv.merge_cells("A2:D2")
r = 4
hdr_row(pv, r, ["Area", "Esercizio", "Dose", "Guarda che…"]); r += 1
prev = [
    ("Ginocchio/anca", "Ponte glutei (bi → mono)", "2×10", "spinta coi talloni, bacino stabile"),
    ("Ginocchio (posteriore coscia)", "Eccentrico ischio leggero (solo U17) / hip hinge (U14)", "2×6", "schiena dritta, scendi lento"),
    ("Caviglia", "Equilibrio monopodalico + calf raise", "2×20″ + 2×12", "piede saldo, sali sulle punte controllata"),
    ("Atterraggio", "Hop & stick / soft landing (già nei blocchi)", "—", "ginocchia in linea, mai a X"),
    ("Spalla", "Circonduzioni + 'spalla a Y/T/W' a corpo libero", "2×10", "scapole ferme, movimento pulito"),
]
for i, row in enumerate(prev):
    rr = r + i
    for c, val in enumerate(row, 1):
        cell = pv.cell(rr, c, val); cell.border = BORD; cell.alignment = LFT if c in (1,2,4) else CTR
for col, w in {"A": 24, "B": 42, "C": 16, "D": 40}.items():
    pv.column_dimensions[col].width = w

# ============================================================ MINI-TEST FINALE
mt = wb.create_sheet("Mini-test finale")
mt.sheet_view.showGridLines = False
section_title(mt, "A1", "MINI-TEST FINALE (fine W3) — checklist per il tecnico")
mt["A2"] = "Non servono strumenti. Segna 🟢 ok / 🟡 migliorabile / 🔴 da rivedere. Segnala i 🔴 e le forti asimmetrie dx/sx al preparatore."
mt["A2"].font = Font(italic=True, size=10, color=GREY)
mt.merge_cells("A1:H1"); mt.merge_cells("A2:H2")
r = 4
hdr_row(mt, r, ["Nome atleta", "Gruppo", "Squat (profondità/ginocchia)", "Equilibrio dx (20-30″)", "Equilibrio sx (20-30″)",
                "Hop&stick dx/sx", "Atterraggio (morbido/silenzioso)", "Plank (qualità)", "Note / segnalazioni"][:9])
r += 1
for i in range(18):
    rr = r + i
    for c in range(1, 10):
        cell = mt.cell(rr, c, ""); cell.border = BORD
widths = [20, 10, 22, 18, 18, 16, 22, 16, 28]
for c, w in enumerate(widths, 1):
    mt.column_dimensions[get_column_letter(c)].width = w

wb.save(OUT)
print("OK ->", OUT, "| fogli:", wb.sheetnames)
